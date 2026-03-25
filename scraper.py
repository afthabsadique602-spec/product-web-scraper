import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Convert rating text -> number
rating_map = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5
}

data = []

print("Starting clean book scraper... Scraping pages 1 to 5.")

# Loop through multiple pages
for page in range(1, 6):
    url = f"https://books.toscrape.com/catalogue/page-{page}.html"
    print(f"Scraping {url}...")
    
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")

    books = soup.find_all("article", class_="product_pod")

    for book in books:
        # Title
        title = book.h3.a["title"]

        # Price (cleaned)
        price_text = book.find("p", class_="price_color").text
        # Sometimes £ comes through as an encoded char like Â£, so replacing both is safer
        price = float(price_text.replace("£", "").replace("Â", ""))

        # Availability
        availability = book.find("p", class_="instock availability").text.strip()

        # Rating (converted)
        rating_text = book.find("p")["class"][1]
        rating = rating_map.get(rating_text, 0)

        # Product link (full URL)
        # On catalogue/page-X.html, links are 'book-url_1/index.html', or '../book-url_1/index.html'
        # Replacing '../' just in case the link resolves relative to the catalogue folder
        link = book.h3.a["href"].replace('../', '')
        full_link = "https://books.toscrape.com/catalogue/" + link

        data.append([title, price, rating, availability, f'=HYPERLINK("{full_link}", "Open Link")'])

# Create DataFrame
df = pd.DataFrame(data, columns=["Title", "Price", "Rating", "Availability", "Link"])

# Save to Excel with professional formatting
excel_filename = "product_data_clean.xlsx"
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Clean Data')
    worksheet = writer.sheets['Clean Data']
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                         
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    for cell in worksheet[1]: 
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    worksheet.row_dimensions[1].height = 25
        
    # Apply currency format to the Price column and text wrapping to all
    for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, max_row=worksheet.max_row):
        for idx, cell in enumerate(row):
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            # Column B (idx==1) is Price
            if idx == 1:
                cell.number_format = '£#,##0.00'
            # Column C (idx==2) is Rating
            if idx == 2:
                cell.alignment = Alignment(horizontal='center', vertical='top')
            # Column E (idx==4) is Link
            if idx == 4:
                cell.font = Font(color="0563C1", underline="single")
            
    for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
        for cell in row:
            cell.border = thin_border
    
    column_widths = {
        'A': 50, # Title
        'B': 15, # Price
        'C': 15, # Rating
        'D': 20, # Availability
        'E': 70  # Link
    }
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
        
    worksheet.freeze_panes = 'A2'

print(f"Scraping completed! Clean data saved to {excel_filename}")
